from model.conta import Conta
from openpyxl import load_workbook

def carrega_banco_de_dados():
    workbook = load_workbook('contas.xlsx')

    # seleciona pagina da planilha
    ws1 = workbook.active

    seed = []
    for numero_da_conta, titular, saldo, cheque_especial, max_cheque_especial, senha in ws1.iter_rows(min_row=1, max_col=6, values_only=True):
        seed.append(Conta(numero_da_conta, titular, saldo, cheque_especial, max_cheque_especial, senha))
    return seed




seed = carrega_banco_de_dados()


def atualiza_banco_de_dados(numero_da_conta, saldo, cheque_especial):
    workbook = load_workbook('contas.xlsx')

    # seleciona pagina da planilha
    ws1 = workbook.active

    ws1.cell(row=numero_da_conta, column=3, value=saldo)

    ws1.cell(row=numero_da_conta, column=4, value=cheque_especial)

    workbook.save('contas.xlsx')

    workbook.close()



def contas_abertas():
    contas_abertas = ''
    for indice, conta in enumerate(seed):
        contas_abertas += "({}) {} \n".format(indice + 1, conta.get_titular())
    return contas_abertas


