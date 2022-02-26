import time

from model.conta import Conta
import random
from seed.conta import seed as SeedConta, contas_abertas, atualiza_banco_de_dados
from openpyxl import load_workbook


def criar_banco_de_dados(numero_da_conta, titular, saldo, cheque_especial_disponivel, max_cheque_especial, senha):
    #carrega planilha
    workbook = load_workbook('contas.xlsx')

    #seleciona pagina da planilha
    ws1 = workbook.active

    #pega dados para salvar
    salvar_conta = [numero_da_conta, titular, saldo, cheque_especial_disponivel, max_cheque_especial, senha]

    #adiciona dados na planilha
    ws1.append(salvar_conta)

    #salva alterações
    workbook.save('contas.xlsx')

    #fechar planilha
    workbook.close()

def mostra_cartao(conta):
    print(" --------------------------------------- ")
    print("|	     CARTAO BANCO NUCLEO		    |")
    print("|					                    |")
    print("| Titular: {}".format(conta.get_titular()))
    print("|					                    |")
    print("| Numero da conta: {}".format(conta.get_numero_da_conta()))
    print("|					                    |")
    print("|					                    |")
    print("|					                    |")
    print(" ---------------------------------------")

def deseja_depositar():
    print("Deseja fazer um deposito inicial?")
    deseja_depositar = int(input("(1) Sim (2) Não"))
    if deseja_depositar == 1:
        valor_depositado = 0 + int(input("Digite o valor que deseja depositar: "))
    else:
        valor_depositado = 0
    return valor_depositado

def exibe_menu_inicial():
    print("(1) Criar conta (2) Acessar conta (3) Informações (4) Sair")

def logo_inicial():
    print("**************************")
    print("Bem vindo ao Banco NUCLEO!")
    print("**************************")

def criar_conta():
    nome = input("Digite seu nome: ").strip().upper()
    numero = len(SeedConta) + 1
    valor_depositado = deseja_depositar()
    max_cheque_especial = valor_depositado + random.randrange(0, 300)
    cheque_especial_disponivel = max_cheque_especial
    senha = input("Digite uma senha: ")

    conta = Conta(numero, nome, valor_depositado, cheque_especial_disponivel, max_cheque_especial, senha)
    SeedConta.append(conta)
    criar_banco_de_dados(numero, nome, valor_depositado, cheque_especial_disponivel, max_cheque_especial, senha)
    print("\n" * 100)
    print("Conta {} criada. \nTitular: {} - Saldo: R$ {} - Cheque especial liberado: R$ {}".format(conta.get_numero_da_conta(), conta.get_titular(), conta.get_saldo(), conta.get_max_cheque_especial()))
    mostra_cartao(conta)
    input("Aperte qualquer tecla para continuar...")
    main()

def acessar_conta():

    workbook = load_workbook('contas.xlsx')

    ws1 = workbook.active

    numero = int(input("Numero da conta: "))
    print(ws1['B{}'.format(numero)].value)
    senha = input("senha: ")

    if senha == ws1['F{}'.format(numero)].value:
        print("Senha correta!")
        print("Acessando...!")
        time.sleep(1)
        print("Acessando..")
        time.sleep(1)
    else:
        print("Numero da conta ou senha incorretos!")
        acessar_conta()

    workbook.save()
    workbook.close()

    usuario_conectado = SeedConta[numero - 1]

    exibir_usuario_saldo(usuario_conectado)
    exibe_menu_da_conta(usuario_conectado)

def exibir_usuario_saldo(usuario_conectado):
    print("Bem vindo, {}".format(usuario_conectado.get_titular()))
    print("Saldo: R$ {}".format(usuario_conectado.get_saldo()))
    print("Saldo + Cheque Especial: RS {}\n".format(usuario_conectado.get_limite()))

def exibe_menu_da_conta(usuario_conectado):
    print("\n" * 100)
    logo_inicial()
    exibir_usuario_saldo(usuario_conectado)

    navega_menu_conta = int(input("(1) Depositar (2) Sacar (3) Transferir (4) Sair"))

    if navega_menu_conta == 1:
        repetir = 0
        while repetir != 2:
            valor = float(input("Quanto deseja depositar?").replace(",", "."))
            usuario_conectado.depositar(valor)
            atualiza_banco_de_dados(usuario_conectado.get_numero_da_conta(), usuario_conectado.get_saldo(),
                                    usuario_conectado.get_cheque_especial())
            print("Saldo atualizado: R$ {}".format(usuario_conectado.extrato()))
            print("Deseja fazer um novo deposito?")
            repetir = int(input("(1) Sim (2) Não"))
            if repetir == 2:
                exibe_menu_da_conta(usuario_conectado)


    elif navega_menu_conta == 2:
        repetir = 0
        while repetir != 2:
            valor = float(input("Quanto deseja sacar?").replace(",", "."))
            usuario_conectado.sacar(valor)
            atualiza_banco_de_dados(usuario_conectado.get_numero_da_conta(), usuario_conectado.get_saldo(),
                                    usuario_conectado.get_cheque_especial())
            print("Saldo atualizado: R$ {}".format(usuario_conectado.extrato()))
            print("Deseja fazer um novo saque?")
            repetir = int(input("(1) Sim (2) Não"))
            if repetir == 2:
                exibe_menu_da_conta(usuario_conectado)

    elif navega_menu_conta == 3:
        repetir = 0
        while repetir != 2:
            valor = float(input("Quanto deseja transferir?").replace(",", "."))
            destino = int(input("Numero da conta destino: ")) - 1
            usuario_conectado.tranferencia(valor, SeedConta[destino])
            atualiza_banco_de_dados(usuario_conectado.get_numero_da_conta(), usuario_conectado.get_saldo(),
                                    usuario_conectado.get_cheque_especial())
            print("Saldo atualizado: R$ {}".format(usuario_conectado.extrato()))
            print("Deseja fazer uma nova transferencia?")
            repetir = int(input("(1) Sim (2) Não"))
            if repetir == 2:
                print("\n" * 100)
                exibe_menu_da_conta(usuario_conectado)

    elif navega_menu_conta == 4:
        atualiza_banco_de_dados(usuario_conectado.get_numero_da_conta(), usuario_conectado.get_saldo(),
                                usuario_conectado.get_cheque_especial())
        print("\n" * 100)
        main()
    else:
        input("Digite um valor válido!")

def main():

    print("\n" * 100)
    logo_inicial()
    exibe_menu_inicial()
    menu_escolha = int(input())

    if menu_escolha == 1:

        criar_conta()

    elif menu_escolha == 2:

        acessar_conta()

    elif menu_escolha == 3:

        main()

    elif menu_escolha == 4:
        print("Até Mais..")
        quit()

    else:

        input("Digite um valor válido!")

if __name__ == "__main__":
    loop = 0
    while loop == 0:
        main()
